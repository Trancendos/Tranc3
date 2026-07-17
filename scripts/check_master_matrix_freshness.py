#!/usr/bin/env python3
"""Fail CI if Trancendos_Master_Service_Matrix.xlsx is stale vs. its generator.

build_master_service_matrix.py derives the workbook from real repo state
(docker-compose.production.yml, workers/*, CSV anchors). Nothing previously
enforced that the committed .xlsx actually matches what the generator would
produce right now, so it could silently drift out of sync with the code it
claims to describe. This script regenerates the workbook into a temp file and
diffs every sheet's cell values against the committed copy, ignoring the
volatile file-level metadata (created/modified timestamps) that openpyxl
stamps on every save. Run it, then `git diff` to see if the drift is real, or
run `python scripts/build_master_service_matrix.py` to refresh it.
"""

from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
COMMITTED = ROOT / "docs" / "architecture" / "ea-workbook" / "Trancendos_Master_Service_Matrix.xlsx"
GENERATOR = ROOT / "scripts" / "build_master_service_matrix.py"


def _sheet_values(path: Path) -> dict[str, list[tuple]]:
    wb = load_workbook(path, data_only=False)
    return {
        name: [tuple(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames
    }


def main() -> int:
    if not COMMITTED.exists():
        print(f"MISSING: {COMMITTED} does not exist yet — run the generator first.")
        return 1

    with tempfile.TemporaryDirectory() as tmp:
        backup = Path(tmp) / "committed_backup.xlsx"
        shutil.copyfile(COMMITTED, backup)

        # The generator writes straight to COMMITTED (it has no --output flag);
        # run it in place, then diff against the pre-run backup, then restore
        # the backup if nothing of substance changed so a clean run leaves the
        # working tree untouched.
        result = subprocess.run(
            [sys.executable, str(GENERATOR)],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            print("FAILED to run the generator for comparison:")
            print(result.stdout)
            print(result.stderr)
            shutil.copyfile(backup, COMMITTED)
            return 1

        committed = _sheet_values(backup)
        regenerated = _sheet_values(COMMITTED)
        backup_bytes = backup.read_bytes()

    if committed.keys() != regenerated.keys():
        print("SHEET MISMATCH")
        print("  committed:  ", sorted(committed.keys()))
        print("  regenerated:", sorted(regenerated.keys()))
        return 1

    stale = False
    for name in committed:
        if committed[name] != regenerated[name]:
            stale = True
            print(f"STALE SHEET: {name!r} differs between committed and regenerated workbook.")
            c_rows, r_rows = committed[name], regenerated[name]
            for i, (c_row, r_row) in enumerate(zip(c_rows, r_rows, strict=False)):
                if c_row != r_row:
                    print(f"  first differing row {i}: committed={c_row!r} regenerated={r_row!r}")
                    break
            if len(c_rows) != len(r_rows):
                print(f"  row count: committed={len(c_rows)} regenerated={len(r_rows)}")

    if stale:
        print(
            "\nTrancendos_Master_Service_Matrix.xlsx is out of date. Run "
            "`python scripts/build_master_service_matrix.py` and commit the result."
        )
        return 1

    # Content is identical; restore the original bytes so a clean run doesn't
    # leave a spurious diff behind from openpyxl's save-time metadata alone.
    COMMITTED.write_bytes(backup_bytes)
    print("OK: Trancendos_Master_Service_Matrix.xlsx matches its generator.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
