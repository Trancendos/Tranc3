#!/usr/bin/env python3
"""Build the Trancendos Master Service Matrix — real data only.

Mirrors the schema of the uploaded "Infinity Services Matrix.xlsx" template
(Services / Routes / Endpoints / Dependencies / Security / Deployment /
Change Log / Deprecations / Executive Dashboard / Governance Checks /
Improvement Roadmap) but every row is populated from facts verified against
the actual Tranc3 codebase during this session's review of PR #221 — no
placeholder/example data adopted from the template.
"""

import datetime
import os
import re

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def scan_workers():
    """Real, structural scan of every workers/* directory — not a deep-dive
    verification like the 8 anchor components above. Records only what can
    be checked mechanically: presence of worker.py/Dockerfile, whether the
    service is registered in docker-compose.production.yml, its real routed
    port there, and whether a /health reference exists in the source."""
    workers_dir = os.path.join(REPO_ROOT, "workers")
    with open(os.path.join(REPO_ROOT, "docker-compose.production.yml")) as f:
        compose_services = yaml.safe_load(f).get("services") or {}

    def compose_port(service_key):
        # Real YAML parse, not a line-based regex — handles both block-list
        # (`ports:\n  - "8000:8000"`) and inline-array (`ports: ["8000:8000"]`)
        # forms, which normalise to the same Python list. Precedence matches
        # scripts/port_registry_validate.py's documented convention (PORT env
        # → Traefik loadbalancer label → published ports), so internal-only
        # services that route via Traefik without publishing a host port
        # (e.g. remotion-render-service) still resolve.
        svc = compose_services.get(service_key)
        if not svc:
            return None

        env = svc.get("environment")
        if isinstance(env, dict) and env.get("PORT") is not None:
            return str(env["PORT"])
        if isinstance(env, list):
            for e in env:
                if str(e).startswith("PORT="):
                    return str(e).split("=", 1)[1]

        for label in svc.get("labels") or []:
            m = re.match(r".*loadbalancer\.server\.port=(\d+)", str(label))
            if m:
                return m.group(1)

        for p in svc.get("ports") or []:
            m = re.match(r"^(?:[\d.]+:)?(\d+):\d+$", str(p).strip('"'))
            if m:
                return m.group(1)
        return None

    def compose_has_service(service_key):
        return service_key in compose_services

    rows = []
    for name in sorted(os.listdir(workers_dir)):
        path = os.path.join(workers_dir, name)
        if not os.path.isdir(path):
            continue
        worker_py = os.path.join(path, "worker.py")
        dockerfile = os.path.join(path, "Dockerfile")
        has_worker_py = os.path.exists(worker_py)
        has_dockerfile = os.path.exists(dockerfile)
        has_health = False
        if has_worker_py:
            with open(worker_py, errors="ignore") as f:
                src = f.read()
            has_health = '"/health"' in src or "'/health'" in src
        in_compose = compose_has_service(name)
        port = compose_port(name) if in_compose else None
        rows.append(
            [
                name,
                has_worker_py,
                has_dockerfile,
                in_compose,
                port or "",
                has_health,
            ]
        )
    return rows


wb = Workbook()

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, header, rows, widths=None, start_row=1):
    for c, h in enumerate(header, start=1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, row=start_row, ncols=len(header))
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    if widths:
        autosize(ws, widths)


# ---------------------------------------------------------------------------
# Overview
# ---------------------------------------------------------------------------
ov = wb.active
ov.title = "Overview"
ov["A1"] = "Trancendos Master Service Matrix"
ov["A1"].font = TITLE_FONT
ov["A3"] = (
    "Real, code-verified service/route/endpoint/dependency/security/deployment "
    "inventory for the 8 real, verified components covered so far. Built from "
    "the same schema as the uploaded Infinity Services Matrix template, but "
    "every value below is checked against actual source in Trancendos/Tranc3 "
    "(commit 182b52b) — no placeholder or example data was carried over."
)
ov["A3"].alignment = WRAP
ov.merge_cells("A3:F3")
ov.row_dimensions[3].height = 60

fields = [
    ("Workbook Name", "Trancendos Master Service Matrix"),
    (
        "Purpose",
        "Real service/route/endpoint/dependency/security/deployment catalogue for verified anchor components",
    ),
    ("Owner", "Andrew Porter (Trancendos)"),
    ("Last Updated", datetime.datetime(2026, 7, 16)),
    (
        "Source of Truth",
        "Verified against actual code: src/, workers/, docker-compose.production.yml, infra/traefik/, deploy/terraform/ — see Notes columns for exact file references",
    ),
    (
        "Scope",
        "8 components deep-verified (Services/Routes/Endpoints/Dependencies/Security/Deployment sheets): The Spark, The Digital Grid, Infinity, The Void, The Workshop, The Observatory, Traefik, health-aggregator. Plus a broad structural scan of all 94 real workers/* directories (see All Workers sheet) — out of ~90 services documented in CLAUDE.md's worker map",
    ),
    (
        "Companion doc",
        "docs/architecture/ea-workbook/ in the Tranc3 repo — this workbook and that CSV workbook describe the same 6 anchor services plus 2 infra components; keep both in sync",
    ),
    (
        "Honesty rule",
        "Every Status/Coverage cell reflects what the code actually does today, including known gaps and broken features — not aspirational state",
    ),
]
r = 5
for k, v in fields:
    ov.cell(row=r, column=1, value=k).font = Font(bold=True)
    ov.cell(row=r, column=2, value=v).alignment = WRAP
    r += 1

ov["A14"] = "Live Architecture Scorecard (real counts, computed from the sheets below)"
ov["A14"].font = Font(bold=True, size=12)
scorecard_header_row = 15
scorecard = [
    (
        "Total components",
        "=COUNTA(Services!A2:A100)",
        "Real components catalogued in this workbook",
    ),
    (
        "Live/Active components",
        '=COUNTIF(Services!H2:H100,"Live")',
        "Status=Live in Services sheet",
    ),
    (
        "Building/migrating components",
        '=COUNTIF(Services!H2:H100,"Building")',
        "In-flight migrations (e.g. The Void: CF Worker to self-hosted)",
    ),
    (
        "Critical components",
        '=COUNTIF(Services!J2:J100,"Critical")',
        "CriticalityCode=CRT-001 equivalents",
    ),
    (
        "Components missing an owner",
        '=COUNTIF(Services!G2:G100,"")',
        "Should be 0 — every real component here has a named Lead AI owner",
    ),
    (
        "Public endpoints",
        '=COUNTIF(Endpoints!J2:J100,"Public")',
        "Endpoints with a real public IP/DNS record (see Endpoints sheet)",
    ),
    (
        "Critical dependencies",
        '=COUNTIF(Dependencies!F2:F100,"Critical")',
        "Hard dependencies whose failure has Critical impact",
    ),
    (
        "Known real defects (see Improvement Roadmap)",
        "=COUNTIF('Improvement Roadmap'!G2:G100,\"Not started\")",
        "Real, verified gaps not yet fixed — not fictional governance noise",
    ),
]
ov.cell(row=scorecard_header_row, column=1, value="Metric").font = Font(bold=True)
ov.cell(row=scorecard_header_row, column=2, value="Value").font = Font(bold=True)
ov.cell(row=scorecard_header_row, column=3, value="Meaning").font = Font(bold=True)
for i, (metric, formula, meaning) in enumerate(scorecard, start=scorecard_header_row + 1):
    ov.cell(row=i, column=1, value=metric)
    ov.cell(row=i, column=2, value=formula)
    ov.cell(row=i, column=3, value=meaning).alignment = WRAP

autosize(ov, [34, 60, 55, 20, 20, 20])

# ---------------------------------------------------------------------------
# Reference Data (kept close to the template's lookup lists, trimmed to what's
# actually used below)
# ---------------------------------------------------------------------------
ref = wb.create_sheet("Reference Data")
ref_rows = [
    (
        "Status Values (real, from docs/architecture/ea-workbook/README.md STS- codes)",
        "Planned, Building, Deploying, Live/Active, Deploying, Failed, Retired",
    ),
    (
        "Environment Values (real)",
        "Local (make dev-api), Production (ENV-005, SRV-CITADEL-01 — the only real production host)",
    ),
    (
        "Criticality Values (real, CRT- codes)",
        "CRT-001 Critical (all 8 components here are CRT-001 — no lower-criticality anchor yet catalogued)",
    ),
    (
        "Auth Values actually implemented",
        "Bearer JWT (HTTPBearer auto_error=False, src/auth/facade.py), OAuth2 authorization_code+PKCE (Infinity only, workers/infinity-auth/router.py), none enforced for Traefik dashboard on internal network",
    ),
    (
        "ACME/TLS (real, infra/traefik/traefik.yml)",
        "certificatesResolvers.letsencrypt.acme.tlsChallenge (TLS-ALPN-01) for external; internal mTLS TLSOption on :8443 for worker-to-worker",
    ),
    (
        "Network (real)",
        "Docker bridge tranc3-net, subnet 172.28.0.0/16 (docker-compose.production.yml)",
    ),
    (
        "Naming pattern actually used",
        "lowercase-hyphenated container_name / service key in docker-compose.production.yml; canonical Trancendos code-name + Tier-3 Lead AI owner per CLAUDE.md",
    ),
]
for i, (a, b) in enumerate(ref_rows, start=2):
    ref.cell(row=i, column=1, value=a).alignment = WRAP
    ref.cell(row=i, column=2, value=b).alignment = WRAP
ref.cell(row=1, column=1, value="Field")
ref.cell(row=1, column=2, value="Real value / source")
style_header(ref, 1, 2)
autosize(ref, [45, 90])

# ---------------------------------------------------------------------------
# Services
# ---------------------------------------------------------------------------
svc = wb.create_sheet("Services")
svc_header = [
    "Service ID",
    "Service Name",
    "Canonical Trancendos Name",
    "Description",
    "Service Type",
    "Business Domain",
    "Owner (Tier 3 Lead AI)",
    "Status",
    "Environment",
    "Criticality",
    "User Facing",
    "Internal Only",
    "Data Class",
    "Real Code Path / Port",
    "Notes",
]
svc_rows = [
    [
        "SRV-001",
        "MCP Tool Registry",
        "The Spark",
        "JSON-RPC 2.0 tool discovery/invocation over HTTP/SSE",
        "API",
        "Platform Core",
        "Norman Hawkins",
        "Live",
        "Production",
        "Critical",
        "No",
        "Yes",
        "Internal",
        "src/mcp/ — mounted on tranc3-backend :8000, routes under /mcp",
        "60s asyncio timeout on tools/call; errors returned as JSON-RPC envelope inside HTTP 200, not 4xx; bearer JWT auth, no OAuth2 scope enforced",
    ],
    [
        "SRV-002",
        "Workflow Engine",
        "The Digital Grid",
        "Topological BFS DAG executor, parallel layers",
        "Backend",
        "Platform Core",
        "Tyler Towncroft",
        "Live",
        "Production",
        "Critical",
        "No",
        "Yes",
        "Internal",
        "src/workflow/, workers/workflow-engine-service/ — :8034, POST /workflows",
        "AutoScalingEnabled=TRUE (max 3) but docker compose --scale can't work yet: fixed container_name + fixed host port 8034:8034 both block it",
    ],
    [
        "SRV-003",
        "Identity Provider",
        "Infinity",
        "OAuth2/OIDC auth, session mgmt, MFA",
        "Auth",
        "Security",
        "The Guardian (Anchor: Orb of Orisis)",
        "Live",
        "Production",
        "Critical",
        "Yes",
        "No",
        "Restricted",
        "workers/infinity-auth/ — :8005, POST /auth/token, GET /auth/verify",
        "authorization_code (+PKCE) works; refresh_token grant is broken (queries nonexistent refresh_tokens table); token_endpoint issues an opaque secrets.token_urlsafe(32) string, not a JWT, so it fails /auth/verify too",
    ],
    [
        "SRV-004",
        "Secrets Vault",
        "The Void",
        "AES-GCM encrypted secrets store, Shamir unseal",
        "Utility",
        "Security",
        "Prometheus",
        "Building",
        "Production",
        "Critical",
        "No",
        "Yes",
        "Restricted",
        "workers/vault-service/ — :8038, GET /secrets/{id}",
        "In-flight migration from CF Worker infinity-void; GET /secrets/{id} returns metadata only (key/tags/ttl/version), never the decrypted value — no endpoint currently exposes it",
    ],
    [
        "SRV-005",
        "Git Hosting & CI/CD",
        "The Workshop",
        "Self-hosted Forgejo git + Actions runner",
        "Utility",
        "DevOps",
        "Larry Lowhammer",
        "Live",
        "Production",
        "Critical",
        "Yes",
        "No",
        "Internal",
        "deploy/forgejo/ — Forgejo container :3000 (Traefik), SSH :2222",
        "Real Traefik rule Host(trancendos.com)&&PathPrefix(/the-workshop); the-workshop.trancendos.com CNAME (real DNS record) 301-redirects to the canonical path via the forgejo-subdomain router",
    ],
    [
        "SRV-006",
        "Audit & Monitoring",
        "The Observatory",
        "Metrics, tracing, audit log aggregation",
        "Backend",
        "Platform Core",
        "Norman Hawkins",
        "Live",
        "Production",
        "Critical",
        "No",
        "Yes",
        "Restricted",
        "workers/observatory/ — :8065, GET /logs, /health",
        "No real persistence: OBSERVATORY_DB_PATH is mounted but never opened; audit events live only in an in-process deque (src/observability/observatory.py) and are lost on restart",
    ],
    [
        "SRV-007",
        "Reverse Proxy / Ingress",
        "Traefik",
        "TLS termination, routing, internal mTLS, ACME",
        "Router",
        "Infrastructure",
        "Cornelius MacIntyre",
        "Live",
        "Production",
        "Critical",
        "Yes",
        "No",
        "Public",
        "docker-compose.production.yml (tranc3-traefik) + infra/traefik/traefik.yml — :80/:443/:8443(mTLS)/:8090(ping)",
        "certResolver=letsencrypt uses tlsChallenge (TLS-ALPN-01); --ping=true fixed this session so the compose healthcheck (traefik healthcheck --ping) actually passes",
    ],
    [
        "SRV-008",
        "Health Aggregator",
        "health-aggregator",
        "Polls every registered service /health every 30s",
        "Utility",
        "Platform Core",
        "Cornelius MacIntyre",
        "Live",
        "Production",
        "Medium",
        "No",
        "Yes",
        "Internal",
        "workers/health-aggregator/ — :8029, POST/GET /services",
        "Hardcoded SERVICE_REGISTRY had 11 P3-block services on wrong ports (alphabetical vs real compose mapping) — fixed this session; scripts/register_ea_workbook_services.py can feed it from the CMDB but nothing calls that script automatically yet",
    ],
]
write_rows(
    svc, svc_header, svc_rows, widths=[10, 22, 20, 42, 12, 16, 26, 10, 12, 10, 10, 10, 12, 46, 70]
)

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
rte = wb.create_sheet("Routes")
rte_header = [
    "Rule ID",
    "Source",
    "Target Service",
    "Request Type",
    "Protocol",
    "Port(s)",
    "Trust Level",
    "Route Action",
    "Real Traefik Rule / Path",
    "Notes",
]
rte_rows = [
    [
        "RTE-001",
        "External Client",
        "Traefik",
        "HTTPS request",
        "HTTPS",
        443,
        "Untrusted",
        "Proxy",
        "entrypoints.websecure, tls.certResolver=letsencrypt",
        "Public internet entrypoint; HTTP :80 redirects to HTTPS per infra/traefik/traefik.yml",
    ],
    [
        "RTE-002",
        "Traefik",
        "The Spark",
        "HTTP (internal)",
        "HTTP",
        8000,
        "Internal",
        "Allow",
        "Mounted on tranc3-backend, no dedicated Traefik router label found for :8000 in compose",
        "The Spark is reached as part of the main backend app, not a standalone Traefik-routed service",
    ],
    [
        "RTE-003",
        "Traefik",
        "The Digital Grid",
        "HTTP (internal)",
        "HTTP",
        8034,
        "Internal",
        "Allow",
        "workflow-engine-service, published host port 8034:8034",
        "Fixed host port blocks any future scale-out beyond 1 replica",
    ],
    [
        "RTE-004",
        "Traefik",
        "Infinity",
        "HTTP (internal)",
        "HTTP",
        8005,
        "Internal",
        "Allow",
        "infinity-auth container",
        "authorization_code + PKCE only; no /oauth2/token, no client_credentials",
    ],
    [
        "RTE-005",
        "Traefik",
        "The Void",
        "HTTP (internal)",
        "HTTP",
        8038,
        "Internal",
        "Allow",
        "vault-service container",
        "Internal hop is plain HTTP — no TLS/mTLS configured for this in-cluster call today",
    ],
    [
        "RTE-006",
        "Traefik",
        "The Workshop",
        "HTTP (internal)",
        "HTTPS",
        3000,
        "Internal",
        "Allow",
        "Host(`trancendos.com`) && PathPrefix(`/the-workshop`), tls.certresolver=letsencrypt",
        "Host=the-workshop.trancendos.com is handled by a separate router (forgejo-subdomain) that 301-redirects to this path — see RTE-006b",
    ],
    [
        "RTE-006b",
        "Traefik",
        "The Workshop",
        "HTTP (internal)",
        "HTTPS (301 redirect)",
        3000,
        "Internal",
        "Allow",
        "Host(`the-workshop.trancendos.com`), middlewares=forgejo-subdomain-redirect, tls.certresolver=letsencrypt",
        "forgejo-subdomain router — 301s to https://trancendos.com/the-workshop/ since Forgejo's ROOT_URL is path-based and would conflict with its own generated links if served directly at this subdomain",
    ],
    [
        "RTE-007",
        "Traefik",
        "The Observatory",
        "HTTP (internal)",
        "HTTP",
        8065,
        "Internal",
        "Allow",
        "observatory container, OBSERVATORY_PORT env (fixed this session)",
        "Dockerfile CMD/HEALTHCHECK previously hardcoded 8040, ignoring compose's OBSERVATORY_PORT=8065 — fixed",
    ],
    [
        "RTE-008",
        "health-aggregator",
        "All 6 anchor services",
        "HTTP GET /health",
        "HTTP",
        "varies",
        "Internal",
        "Allow",
        "SERVICE_REGISTRY in workers/health-aggregator/worker.py",
        "Polls every 30s; port mapping for the P3 block was wrong for 11 services, fixed this session",
    ],
    [
        "RTE-009",
        "Worker-to-worker (any)",
        "Any internal worker",
        "mTLS",
        "HTTPS",
        8443,
        "Internal (mTLS)",
        "Allow",
        "entryPoints.internal in infra/traefik/traefik.yml, TLSOption requires client cert signed by internal-ca.pem",
        "Separate internal-only entrypoint from the public :443 — not exposed externally",
    ],
]
write_rows(rte, rte_header, rte_rows, widths=[10, 20, 20, 18, 10, 10, 14, 12, 55, 65])

# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
ep = wb.create_sheet("Endpoints")
ep_header = [
    "Endpoint ID",
    "Service Name",
    "Hostname / DNS record",
    "Port",
    "Protocol",
    "Interface Type",
    "TLS",
    "Certificate Source",
    "Public Exposure?",
    "Notes",
]
ep_rows = [
    [
        "EP-001",
        "Traefik (apex)",
        "trancendos.com (A, cloudflare_record.apex)",
        443,
        "HTTPS",
        "Public",
        "Yes",
        "Let's Encrypt via tlsChallenge",
        "Public",
        "Real Terraform record (deploy/terraform/oci-citadel-dns.tf); resolves to oci_core_public_ip.citadel.ip_address at apply time; proxied=false (grey cloud, required for TLS-ALPN-01)",
    ],
    [
        "EP-002",
        "Traefik (www)",
        "www.trancendos.com (A, cloudflare_record.www)",
        443,
        "HTTPS",
        "Public",
        "Yes",
        "Let's Encrypt via tlsChallenge",
        "Public",
        "Real Terraform record; proxied=false",
    ],
    [
        "EP-003",
        "Traefik (api)",
        "api.trancendos.com (A, cloudflare_record.api)",
        443,
        "HTTPS",
        "Public",
        "Yes",
        "Let's Encrypt via tlsChallenge",
        "Public",
        "A record inside the trancendos.com zone — NOT a separate delegated zone; proxied=false",
    ],
    [
        "EP-004",
        "The Workshop / Forgejo",
        "the-workshop.trancendos.com (CNAME, cloudflare_record.the_workshop)",
        "n/a",
        "HTTP(S)",
        "Public (DNS only)",
        "n/a",
        "n/a",
        "Public DNS, 301-redirects to canonical path",
        "Real DNS record provisioned by Terraform; the forgejo-subdomain Traefik router matches this Host header and 301-redirects to https://trancendos.com/the-workshop/",
    ],
    [
        "EP-005",
        "The Workshop / Forgejo (SSH)",
        "trancendos.com",
        2222,
        "SSH",
        "Public",
        "n/a",
        "n/a",
        "Public",
        "git+ssh clone; published as 2222:22 on the Forgejo container itself, not a separate runner node",
    ],
    [
        "EP-006",
        "Traefik dashboard",
        "traefik.trancendos.local",
        8888,
        "HTTP",
        "Internal",
        "No",
        "None",
        "Internal",
        "Local-only Host rule (.local); also published on host port 8888 — not intended for public exposure",
    ],
    [
        "EP-007",
        "Internal mTLS entrypoint",
        "(container network only)",
        8443,
        "HTTPS (mTLS)",
        "Internal",
        "Yes",
        "Internal CA (infra/traefik/certs/internal-ca.pem)",
        "Internal only",
        "Worker-to-worker calls requiring a client cert signed by the internal CA",
    ],
    [
        "EP-008",
        "Traefik ping",
        "(container network only)",
        8090,
        "HTTP",
        "Internal",
        "No",
        "n/a",
        "Internal only",
        "ping.entryPoint per infra/traefik/traefik.yml; compose healthcheck instead uses --ping CLI probe on the default entrypoint",
    ],
]
write_rows(ep, ep_header, ep_rows, widths=[10, 22, 46, 8, 14, 16, 8, 30, 18, 70])

# ---------------------------------------------------------------------------
# Dependencies
# ---------------------------------------------------------------------------
dep = wb.create_sheet("Dependencies")
dep_header = [
    "Dependency ID",
    "Service",
    "Depends On",
    "Dependency Type",
    "Required",
    "Failure Impact",
    "Recovery Strategy (real)",
    "Notes",
]
dep_rows = [
    [
        "DEP-001",
        "The Spark",
        "The Void",
        "Secrets",
        "Yes",
        "Critical",
        "None automated — restart after Void recovers",
        "Cannot resolve secrets needed for tool execution if Void is down",
    ],
    [
        "DEP-002",
        "The Spark",
        "Infinity",
        "Auth",
        "Yes",
        "Critical",
        "None automated",
        "get_current_user dependency requires Infinity-issued JWTs to validate",
    ],
    [
        "DEP-003",
        "The Digital Grid",
        "The Void",
        "Secrets",
        "Yes",
        "Critical",
        "None automated",
        None,
    ],
    ["DEP-004", "The Digital Grid", "Infinity", "Auth", "Yes", "Critical", "None automated", None],
    [
        "DEP-005",
        "The Digital Grid",
        "The Spark",
        "Tool invocation",
        "No (soft)",
        "Medium",
        "Workflow steps that call MCP tools degrade gracefully",
        "Only workflows with MCP tool steps are affected",
    ],
    [
        "DEP-006",
        "Infinity",
        "The Void",
        "Secrets",
        "Yes",
        "Critical",
        "None automated",
        "Infinity cannot sign tokens if it can't reach the vault for its signing secret",
    ],
    [
        "DEP-007",
        "The Void",
        "Infinity",
        "Identity verification",
        "Yes",
        "Critical",
        "None automated",
        "Caller identity verification before secret release",
    ],
    ["DEP-008", "The Workshop", "Infinity", "Auth", "Yes", "High", "None automated", None],
    ["DEP-009", "The Workshop", "The Void", "Secrets", "Yes", "High", "None automated", None],
    ["DEP-010", "The Observatory", "Infinity", "Auth", "Yes", "Critical", "None automated", None],
    [
        "DEP-011",
        "The Observatory",
        "(no real DB dependency)",
        "n/a",
        "No",
        "n/a",
        "n/a",
        "10_databases.csv DB-OBS-001 was previously listed as a hard dependency — corrected this session: the worker never opens OBSERVATORY_DB_PATH, so it cannot cause a real failure",
    ],
    [
        "DEP-012",
        "health-aggregator",
        "All 6 anchor services",
        "Health polling",
        "No",
        "Low",
        "Marks a service unhealthy in its own DB; does not trigger automated remediation",
        "Read-only dependency — a target being down doesn't break health-aggregator itself",
    ],
]
write_rows(dep, dep_header, dep_rows, widths=[12, 20, 24, 16, 12, 12, 46, 60])

# ---------------------------------------------------------------------------
# Security
# ---------------------------------------------------------------------------
sec = wb.create_sheet("Security")
sec_header = [
    "Service Name",
    "Data Classification",
    "Authentication (real)",
    "Authorization (real)",
    "Encryption in Transit",
    "Encryption at Rest",
    "Audit Logging",
    "Notes",
]
sec_rows = [
    [
        "The Spark",
        "Internal",
        "Bearer JWT via Depends(get_current_user), auto_error=False so 401 (not 403) on missing/invalid creds",
        "No OAuth2 scope declared or enforced despite security scheme listing scopes",
        "Internal HTTP only (no TLS on the tranc3-backend hop)",
        "N/A (no persistent store)",
        "Via Observatory audit middleware",
        "src/auth/facade.py:get_current_user_dep raises 401 uniformly for missing and invalid tokens",
    ],
    [
        "The Digital Grid",
        "Internal",
        "Bearer JWT (same facade)",
        "internal-auth only via require_internal_auth",
        "Internal HTTP only",
        "SQLite file, not encrypted at rest",
        "Via Observatory audit middleware",
        None,
    ],
    [
        "Infinity",
        "Restricted",
        "OIDC-style authorization_code+PKCE (workers/infinity-auth); refresh_token grant broken (missing table)",
        "Own JWT issuance/verification; get_current_user requires JWT",
        "External: Traefik TLS (tlsChallenge/TLS-ALPN-01); internal hop plain HTTP",
        "SQLite file, path/volume-mount mismatch means data isn't durably backed up",
        "Via Observatory audit middleware",
        "authorization_code token is opaque (secrets.token_urlsafe(32)), NOT a JWT — /auth/verify and every other protected route on this service will reject it with 401",
    ],
    [
        "The Void",
        "Restricted",
        "internal-auth only",
        "Caller identity check against Infinity before secret release",
        "Internal hop is plain HTTP — no TLS/mTLS configured for this specific in-cluster call",
        "AES-GCM encrypted values; Shamir-split unseal keys, never stored alongside the encrypted DB",
        "Via Observatory audit middleware",
        "GET /secrets/{id} returns metadata only — no endpoint currently returns the decrypted value",
    ],
    [
        "The Workshop",
        "Internal",
        "Forgejo's own auth (registration disabled, sign-in required to view)",
        "Forgejo RBAC",
        "External: Traefik TLS (tlsChallenge)",
        "SQLite (Forgejo's own gitea.db)",
        "Forgejo's own audit/activity log",
        "SSH clone bypasses HTTP auth entirely — key-based, published at host port 2222",
    ],
    [
        "The Observatory",
        "Restricted",
        "Bearer JWT (same facade) for query endpoints",
        "internal-auth",
        "Internal HTTP only",
        "None — in-memory deque only, lost on restart",
        "Self-referential — Observatory IS the audit log, but not durably",
        "OBSERVATORY_DB_PATH configured but never opened; treat any retention/compliance claim about this data surviving a restart as aspirational",
    ],
    [
        "Traefik",
        "Public-facing",
        "None for the ingress itself; per-router labels",
        "N/A (routing layer)",
        "External TLS via tlsChallenge; internal mTLS TLSOption on :8443 requiring client certs signed by internal-ca.pem",
        "N/A",
        "accessLog.format=json with header allowlist (X-Request-ID, X-Forwarded-For, X-Device-Posture kept)",
        "Dashboard (api.dashboard) enabled but insecure=false and only reachable via a .local Host rule — not exposed publicly by design",
    ],
    [
        "health-aggregator",
        "Internal",
        "X-Internal-Secret header (INTERNAL_SECRET env, optional)",
        "n/a",
        "Internal HTTP only",
        "SQLite (health check history)",
        "Own SQLite history, not routed through Observatory",
        "If INTERNAL_SECRET is unset, /services management endpoints are unauthenticated on the internal network",
    ],
]
write_rows(sec, sec_header, sec_rows, widths=[20, 16, 55, 40, 45, 45, 30, 70])

# ---------------------------------------------------------------------------
# Deployment
# ---------------------------------------------------------------------------
dpl = wb.create_sheet("Deployment")
dpl_header = [
    "Service Name",
    "Real Code Path",
    "Runtime",
    "Deployment Method (real)",
    "Rollback Method (real)",
    "Health Check",
    "Last Verified Commit (this session)",
    "Notes",
]
dpl_rows = [
    [
        "The Spark",
        "src/mcp/",
        "Python 3.11 / FastAPI 0.136.3",
        "Part of tranc3-backend — Fly.io (legacy) or manual docker compose",
        "git checkout <sha> + docker compose up -d --build (manual, no automation)",
        "/mcp/health",
        "93d87d3e (2026-06-30)",
        "No dedicated CI artifact/pipeline run tracked yet for this path — 06_service_deployments.csv leaves those columns blank rather than fabricated",
    ],
    [
        "The Digital Grid",
        "src/workflow/, workers/workflow-engine-service/",
        "Python 3.11 / FastAPI 0.136.3",
        "Manual docker compose up -d --build",
        "Same manual pattern",
        "/health",
        "df4ed5e5 (2026-06-25)",
        None,
    ],
    [
        "Infinity",
        "workers/infinity-auth/",
        "Python 3.11 / FastAPI 0.136.3",
        "Manual docker compose up -d --build",
        "Same manual pattern",
        "/health",
        "a3df9128 (2026-07-12)",
        None,
    ],
    [
        "The Void",
        "workers/vault-service/",
        "Python 3.11 / FastAPI 0.136.3",
        "Manual docker compose up -d --build (in-flight CF Worker migration)",
        "Same manual pattern",
        "/health",
        "d06f3b13 (2026-06-29)",
        "Migrating from cloudflare/infinity-void/",
    ],
    [
        "The Workshop",
        "deploy/forgejo/",
        "Forgejo (Go), image codeberg.org/forgejo/forgejo:7",
        "deploy/forgejo/setup.sh + docker compose",
        "deploy/forgejo/recover.sh (topology-aware, real dedicated runbook)",
        "wget http://localhost:3000/-/health",
        "8fbc80b8 (2026-07-12)",
        "The only anchor service with a genuinely dedicated recovery runbook (deploy/forgejo/RUNBOOK.md)",
    ],
    [
        "The Observatory",
        "workers/observatory/",
        "Python 3.11 / FastAPI 0.136.3",
        "Manual docker compose up -d --build",
        "Same manual pattern",
        "/health",
        "d06f3b13 (2026-06-29)",
        "Dockerfile CMD/HEALTHCHECK fixed this session to honor OBSERVATORY_PORT instead of hardcoding 8040",
    ],
    [
        "Traefik",
        "docker-compose.production.yml, infra/traefik/traefik.yml",
        "traefik:v3.0 image",
        "docker compose (part of core infra stack)",
        "docker compose restart traefik / redeploy image tag",
        "traefik healthcheck --ping (CLI probe, not an HTTP path)",
        "--ping=true added this session (182b52b lineage)",
        "Compose healthcheck was silently failing before this fix — --ping was never enabled",
    ],
    [
        "health-aggregator",
        "workers/health-aggregator/",
        "Python 3.11 / FastAPI 0.136.3",
        "Manual docker compose up -d --build",
        "Same manual pattern",
        "/health",
        "0673555 (2026-07-16, this session)",
        "SERVICE_REGISTRY port fix and scripts/post_deploy_verify.py fix landed together",
    ],
]
write_rows(dpl, dpl_header, dpl_rows, widths=[18, 38, 26, 42, 40, 32, 26, 65])

# ---------------------------------------------------------------------------
# Change Log — real commits from this session, not fictional entries
# ---------------------------------------------------------------------------
chg = wb.create_sheet("Change Log")
chg_header = ["Change ID", "Date (UTC)", "Commit", "Summary", "Real Impact"]
chg_rows = [
    [
        "CHG-001",
        "2026-07-16T18:29:43",
        "5c18133",
        "Fixed remaining CMDB/docs bugs found in cubic-dev-ai review batch",
        "Corrected API routes, engine types, network subnet, Forgejo SSH port, Traefik health mechanism across the EA workbook",
    ],
    [
        "CHG-002",
        "2026-07-16T18:43:35",
        "7b346f5",
        "Fixed the real Traefik ping bug + 15 more CMDB findings",
        "Added --ping=true to Traefik's command args — the compose healthcheck was permanently failing before this",
    ],
    [
        "CHG-003",
        "2026-07-16T18:53:04",
        "01efe9e",
        "Fixed health-aggregator port drift; wired EA workbook into the live system",
        "Corrected 11 P3-block services on wrong ports in health-aggregator's SERVICE_REGISTRY; added CI/pre-commit CSV validation; added register_ea_workbook_services.py",
    ],
    [
        "CHG-004",
        "2026-07-16T18:57:24",
        "e9891aa",
        "Fixed deployment chronology",
        "6 anchor deployment rows had DeploymentDate before their own commit's real timestamp — corrected using git show",
    ],
    [
        "CHG-005",
        "2026-07-16T19:13:17",
        "0673555",
        "Fixed real DNS gap + second port-drift bug + script robustness",
        "Found the identical port-mapping bug duplicated in scripts/post_deploy_verify.py; found the workbook was missing 2 of 4 real Terraform-provisioned DNS records",
    ],
    [
        "CHG-006",
        "2026-07-16T23:52:20",
        "182b52b",
        "Fixed ACME challenge type + the-workshop CNAME routing claim",
        "Corrected HTTP-01 to the real tlsChallenge/TLS-ALPN-01; discovered the-workshop CNAME has no matching Traefik host rule",
    ],
    [
        "CHG-007",
        "2026-07-17",
        "pending",
        "Fixed cubic-dev-ai review batch: the register/verify loop never actually worked",
        (
            "health-aggregator's SERVICE_REGISTRY built health_url from 'localhost', which "
            "inside its own bridge-networked container only ever reached itself — every "
            "static check silently probed as down since deployment. Dynamically-registered "
            "services (POST /services) were never included in the poll loop, /status, "
            "/predict, or /metrics at all. register_ea_workbook_services.py wrote loopback "
            "URLs unusable from health-aggregator's container, and deploy-self-hosted.yml's "
            "verify/register steps ran from act-runner, which isn't on tranc3-net and can't "
            "reach compose-published ports via 'localhost' either. Fixed: health-aggregator "
            "now resolves every target (static + dynamic) via Docker service DNS; "
            "register/verify scripts default to compose service names instead of localhost; "
            "act-runner joins tranc3-net (now an explicitly-named external network) so those "
            "CI steps can actually reach the fleet. Also fixed: build_master_service_matrix.py "
            "used a line-regex for compose ports that missed every inline-array `ports: [...]` "
            "service (silently blanked ~15 P3 workers' ports); openpyxl was an undeclared "
            "dependency; Workshop route/endpoint rows still described the pre-redirect-fix "
            "state; Governance Checks' Current Result column was never populated."
        ),
    ],
]
write_rows(chg, chg_header, chg_rows, widths=[10, 20, 10, 55, 75])

# ---------------------------------------------------------------------------
# Deprecations — real, in-flight
# ---------------------------------------------------------------------------
dep2 = wb.create_sheet("Deprecations")
dep2_header = ["Service Name", "Replaced By", "Reason", "Data Migration Done", "Notes"]
dep2_rows = [
    [
        "infinity-void (Cloudflare Worker)",
        "The Void / vault-service (self-hosted)",
        "Zero-cost self-hosted architecture migration (CLAUDE.md)",
        "In Progress",
        "cloudflare/infinity-void/ still exists in-repo as the legacy source; workers/vault-service/ is the real self-hosted replacement, StatusCode=Building",
    ],
    [
        "tranc3-api-gateway (Cloudflare Worker)",
        "Traefik (self-hosted)",
        "Same zero-cost migration",
        "In Progress",
        "cloudflare/trancendos-api-gateway/ still exists; api.trancendos.com now routes through Traefik per Terraform",
    ],
]
write_rows(dep2, dep2_header, dep2_rows, widths=[32, 32, 45, 20, 65])

# ---------------------------------------------------------------------------
# Executive Dashboard — real, formula-driven
# ---------------------------------------------------------------------------
xd = wb.create_sheet("Executive Dashboard")
xd["A1"] = "Executive Dashboard — Trancendos Master Service Matrix"
xd["A1"].font = TITLE_FONT
xd["A3"] = "Component Status Summary"
xd["A3"].font = Font(bold=True)
xd_header_row = 4
xd.cell(row=xd_header_row, column=1, value="Status")
xd.cell(row=xd_header_row, column=2, value="Count")
style_header(xd, xd_header_row, 2)
statuses = ["Live", "Building", "Deploying", "Planned", "Failed", "Retired"]
for i, s in enumerate(statuses, start=xd_header_row + 1):
    xd.cell(row=i, column=1, value=s)
    xd.cell(row=i, column=2, value=f'=COUNTIF(Services!H2:H100,"{s}")')

xd["D3"] = "Dependencies by Failure Impact"
xd["D3"].font = Font(bold=True)
xd.cell(row=xd_header_row, column=4, value="Impact")
xd.cell(row=xd_header_row, column=5, value="Count")
style_header(xd, xd_header_row, 5)
impacts = ["Critical", "High", "Medium", "Low"]
for i, imp in enumerate(impacts, start=xd_header_row + 1):
    xd.cell(row=i, column=4, value=imp)
    xd.cell(row=i, column=5, value=f'=COUNTIF(Dependencies!F2:F100,"{imp}")')

xd["A12"] = "Executive Action Summary (real, from Improvement Roadmap)"
xd["A12"].font = Font(bold=True, size=12)
xd_action_header = 13
xd.cell(row=xd_action_header, column=1, value="Metric")
xd.cell(row=xd_action_header, column=2, value="Value")
xd.cell(row=xd_action_header, column=3, value="Interpretation")
style_header(xd, xd_action_header, 3)
actions = [
    (
        "Real open defects (P1)",
        "=COUNTIFS('Improvement Roadmap'!F2:F100,\"P1\",'Improvement Roadmap'!G2:G100,\"Not started\")",
        "Highest-priority real gaps — fix before claiming full production readiness",
    ),
    (
        "Real open defects (P2)",
        "=COUNTIFS('Improvement Roadmap'!F2:F100,\"P2\",'Improvement Roadmap'!G2:G100,\"Not started\")",
        "Should be scheduled but not urgent",
    ),
    (
        "Components covered",
        "=COUNTA(Services!A2:A100)",
        "Out of ~90 real services documented in CLAUDE.md's worker map",
    ),
    (
        "Deployment automation gap",
        "=COUNTIF('Improvement Roadmap'!A2:A100,\"Automation\")",
        "Manual docker compose deploys remain the norm — see Deployment sheet",
    ),
]
for i, (m, formula, interp) in enumerate(actions, start=xd_action_header + 1):
    xd.cell(row=i, column=1, value=m)
    xd.cell(row=i, column=2, value=formula)
    xd.cell(row=i, column=3, value=interp).alignment = WRAP
autosize(xd, [30, 40, 55, 12, 14])

# ---------------------------------------------------------------------------
# Improvement Roadmap — real, discovered gaps (not fictional governance items)
# ---------------------------------------------------------------------------
road = wb.create_sheet("Improvement Roadmap")
road_header = [
    "Theme",
    "Action",
    "Why it matters",
    "How to complete",
    "Owner",
    "Priority",
    "Status",
]
road_rows = [
    [
        "Routing",
        "Add a Traefik host rule for the-workshop.trancendos.com",
        "Real DNS CNAME is provisioned by Terraform but was unreachable — Forgejo's only rule required Host=trancendos.com",
        "Added a redirect router (forgejo-subdomain, Host(`the-workshop.trancendos.com`)) that 301s to the canonical https://trancendos.com/the-workshop/ path — Forgejo's ROOT_URL is path-based, so serving it directly at the bare subdomain would have conflicted with its own generated links",
        "Larry Lowhammer",
        "P1",
        "Fixed",
    ],
    [
        "Auth",
        "Fix or replace Infinity's authorization_code token contract",
        "The issued opaque token cannot be validated by any of this service's own protected routes (/auth/verify, /auth/me) — a real, functional gap, not a doc error",
        "Either issue a signed JWT from the authorization_code grant too, or add a dedicated validation path for opaque tokens",
        "The Guardian (Anchor: Orb of Orisis)",
        "P1",
        "Not started",
    ],
    [
        "Auth",
        "Fix Infinity's refresh_token grant",
        "The refresh path queries a refresh_tokens table that is never created — refresh currently always fails",
        "Create the missing table/migration, or remove refresh_token support from the advertised grant types until it's implemented",
        "The Guardian (Anchor: Orb of Orisis)",
        "P1",
        "Not started",
    ],
    [
        "Observability",
        "Give The Observatory real persistence",
        "OBSERVATORY_DB_PATH is configured but never opened; all audit/metric data lives only in an in-process deque and is lost on every restart",
        "Wire the configured DB path into an actual SQLite writer in workers/observatory/service.py",
        "Norman Hawkins",
        "P1",
        "Not started",
    ],
    [
        "Automation",
        "Wire scripts/register_ea_workbook_services.py and scripts/post_deploy_verify.py into a real deploy step",
        "Both scripts exist and work (verified this session) but nothing in .forgejo/workflows/ calls either one — the CMDB doesn't drive live monitoring yet",
        "Added a verify-and-register-self-hosted job to deploy-self-hosted.yml that runs both scripts (--soft / continue-on-error, since the underlying Compose deploy step below still doesn't exist)",
        "Cornelius MacIntyre",
        "P2",
        "Fixed",
    ],
    [
        "Automation",
        "Make deploy-self-hosted.yml actually deploy workers/* via Docker Compose",
        "It triggers on workers/** changes but only runs flyctl for trancendos-backend/trancendos-bots — no docker compose up step exists for the self-hosted workers",
        "Add a real Compose deploy step (or document the manual host-level step explicitly as the real process)",
        "Cornelius MacIntyre",
        "P2",
        "Not started",
    ],
    [
        "Scaling",
        "Fix The Digital Grid's autoscaling",
        "AutoScalingEnabled=TRUE (max 3) but both the fixed container_name and the fixed 8034:8034 host port publish block more than one replica today",
        "Remove the fixed container_name and route traffic to replicas via Traefik service discovery instead of the published host port",
        "Tyler Towncroft",
        "P2",
        "Not started",
    ],
    [
        "Coverage",
        "Expand this matrix beyond the 8 deep-verified components",
        "CLAUDE.md documents ~90 real services; this workbook and the EA CSV workbook both cover only 6 anchors + 2 infra components at full verification depth",
        "Add one verified row per additional service across all sheets here, following the same code-verification standard used for the anchors — the All Workers (Broad Scan) sheet is a structural first pass, not a substitute for that depth",
        "Andrew Porter",
        "P3",
        "In progress",
    ],
    [
        "Architecture",
        "Resolve the duplicate 'Void' implementation",
        "workers/the-void/ (289 lines, includes a Rust crypto extension, last touched in commit 634962c) implements the same AES-GCM vault concept as workers/vault-service/ (803 lines, the one actually wired into docker-compose.production.yml and documented throughout this workbook) — but the-void/ is not registered in compose at all and is not deployed anywhere",
        "A human needs to decide: is the-void/ an abandoned earlier attempt safe to remove, or a newer implementation meant to eventually replace vault-service/? Do not delete either without that decision",
        "Prometheus",
        "P2",
        "Not started",
    ],
    [
        "Governance",
        "Merge PR #221",
        "This entire matrix, the EA CSV workbook, and all the real bugfixes found this session live on an open, unmerged PR",
        "Get CI green and merge claude/platform-context-review-ieg1rn to main",
        "Andrew Porter",
        "P2",
        "Not started",
    ],
]
write_rows(road, road_header, road_rows, widths=[16, 46, 60, 60, 30, 10, 14])

# ---------------------------------------------------------------------------
# All Workers (Broad Scan) — every workers/* directory, structurally scanned
# ---------------------------------------------------------------------------
scan = wb.create_sheet("All Workers (Broad Scan)")
scan["A1"] = "All Workers — Broad Structural Scan (not deep-verified like the 8 anchors above)"
scan["A1"].font = Font(bold=True, size=12)
scan["A2"] = (
    "Mechanically checked against real files: presence of worker.py/Dockerfile, "
    "registration in docker-compose.production.yml, real routed port there, and "
    "whether a /health string appears in the source. This is a structural first "
    "pass across all real workers/* directories — treat rows here as a starting "
    "point for the same deep-verification the 8 anchor components received, not "
    "as equivalent confidence."
)
scan["A2"].alignment = WRAP
scan.merge_cells("A2:F2")
scan.row_dimensions[2].height = 45

scan_header_row = 4
scan_header = [
    "Worker Directory",
    "Has worker.py",
    "Has Dockerfile",
    "In docker-compose.production.yml",
    "Real Compose Port",
    "Has /health Reference",
]
for c, h in enumerate(scan_header, start=1):
    scan.cell(row=scan_header_row, column=c, value=h)
style_header(scan, scan_header_row, len(scan_header))

worker_rows = scan_workers()
for r, row in enumerate(worker_rows, start=scan_header_row + 1):
    for c, val in enumerate(row, start=1):
        scan.cell(row=r, column=c, value=val)
scan.freeze_panes = scan.cell(row=scan_header_row + 1, column=1).coordinate
autosize(scan, [32, 14, 14, 26, 16, 18])

scan_summary_row = scan_header_row + len(worker_rows) + 3
scan.cell(row=scan_summary_row, column=1, value="Summary").font = Font(bold=True)
summary_stats = [
    (
        "Total worker directories found",
        f"=COUNTA(A{scan_header_row + 1}:A{scan_header_row + len(worker_rows)})",
    ),
    (
        "Registered in docker-compose.production.yml",
        f"=COUNTIF(D{scan_header_row + 1}:D{scan_header_row + len(worker_rows)},TRUE)",
    ),
    (
        "NOT registered (code exists, not deployed)",
        f"=COUNTIF(D{scan_header_row + 1}:D{scan_header_row + len(worker_rows)},FALSE)",
    ),
    (
        "Has a /health reference in source",
        f"=COUNTIF(F{scan_header_row + 1}:F{scan_header_row + len(worker_rows)},TRUE)",
    ),
]
for i, (label, formula) in enumerate(summary_stats, start=scan_summary_row + 1):
    scan.cell(row=i, column=1, value=label)
    scan.cell(row=i, column=2, value=formula)

# ---------------------------------------------------------------------------
# Governance Checks — real, formula-driven, self-auditing
# ---------------------------------------------------------------------------
gov = wb.create_sheet("Governance Checks")
gov["A1"] = "Governance Checks (live formulas — recalculate in Excel/LibreOffice)"
gov["A1"].font = Font(bold=True, size=12)
gov_header_row = 3
gov_header = ["Check", "Live Formula", "Current Result", "Pass Target"]
for c, h in enumerate(gov_header, start=1):
    gov.cell(row=gov_header_row, column=c, value=h)
style_header(gov, gov_header_row, 4)
checks = [
    ("Every component has a named owner", '=COUNTIF(Services!G2:G100,"")', 0),
    (
        # Per-name match, not an aggregate row-count diff: a count-only
        # comparison would read PASS even if Security's rows didn't
        # correspond to the same components as Services' rows.
        "Every component has a Security row",
        '=SUMPRODUCT((Services!B2:B100<>"")*ISNA(MATCH(Services!B2:B100,Security!A2:A100,0)))',
        0,
    ),
    (
        "Every component has a Deployment row",
        '=SUMPRODUCT((Services!B2:B100<>"")*ISNA(MATCH(Services!B2:B100,Deployment!A2:A100,0)))',
        0,
    ),
    (
        "Critical dependencies have a recovery strategy filled in",
        '=COUNTIFS(Dependencies!F2:F100,"Critical",Dependencies!G2:G100,"")',
        0,
    ),
    (
        "P1 real defects still open",
        "=COUNTIFS('Improvement Roadmap'!F2:F100,\"P1\",'Improvement Roadmap'!G2:G100,\"Not started\")",
        0,
    ),
    (
        "Public endpoints without TLS",
        '=COUNTIFS(Endpoints!I2:I100,"Public",Endpoints!G2:G100,"No")',
        0,
    ),
]
for i, (check, formula, target) in enumerate(checks, start=gov_header_row + 1):
    gov.cell(row=i, column=1, value=check).alignment = WRAP
    gov.cell(row=i, column=2, value=formula)
    gov.cell(row=i, column=3, value=f'=IF(B{i}=D{i},"PASS","FAIL")')
    gov.cell(row=i, column=4, value=target)
autosize(gov, [55, 60, 16, 14])

order = [
    "Overview",
    "Reference Data",
    "Services",
    "Routes",
    "Endpoints",
    "Dependencies",
    "Security",
    "Deployment",
    "Change Log",
    "Deprecations",
    "Executive Dashboard",
    "All Workers (Broad Scan)",
    "Improvement Roadmap",
    "Governance Checks",
]
wb._sheets = [wb[name] for name in order]

out_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "docs",
    "architecture",
    "ea-workbook",
    "Trancendos_Master_Service_Matrix.xlsx",
)
wb.save(out_path)
print("Saved:", out_path)
